"""
excelio.py — Excel reading and writing helpers.

read_workbook_rows : load an uploaded .xlsx into a list-of-lists.
build_export       : write Zee-format show rows to a date-grouped .xlsx,
                     matching the Output_Final.xlsx layout:
                       - row 1: a "Break N" group header over each Ad ST/ET pair
                       - row 2: the 68 column headers (amber, bold)
                       - segment cells shaded light blue
                       - blank separator rows between dates
"""

import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from transform import (OUTPUT_COLS, SHOW_INFO_COLS, BREAK_COLS, SEG_COLS,
                       display_label)

# Shared colours (also used by the web UI).
SEGMENT_BLUE = "FFCFE7F5"     # light blue for segment cells
HEADER_AMBER = "FFFFC000"     # amber header band (matches Output_Final.xlsx)


def read_workbook_rows(file_stream):
    """Read the first sheet of an .xlsx stream into a list of row-lists."""
    wb = load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _mins_missing_fill(value):
    """Heat shading for Mins Missing (bigger gap -> darker red)."""
    try:
        v = abs(float(value))
    except (ValueError, TypeError):
        return None
    if v <= 0:
        return None
    if v < 5:
        color = "FFF0CCCC"
    elif v < 15:
        color = "FFE49999"
    elif v < 30:
        color = "FFD96666"
    else:
        color = "FFCC3333"
    return PatternFill("solid", start_color=color)


def build_export(show_rows, fmt="csv"):
    """
    Build a single-table date-grouped export. Returns (bytes, mimetype, ext).
    Used for the CSV export (CSV cannot hold multiple sheets).
    """
    if fmt == "csv":
        return _build_csv(show_rows)
    # single-sheet xlsx fallback
    wb = Workbook()
    _fill_sheet(wb.active, "Zee TV", show_rows)
    return _save(wb)


def build_multi_sheet_export(all_rows, channel_sheets):
    """
    Build ONE .xlsx workbook with multiple sheets:
      * an "All" sheet containing every show row, plus
      * one sheet per channel.

    Parameters
    ----------
    all_rows : list of show-row dicts (already date-sorted)
    channel_sheets : list of (sheet_name, rows) tuples, in display order

    Returns (bytes, mimetype, extension).
    """
    wb = Workbook()
    # First sheet -> "All"
    _fill_sheet(wb.active, "All", all_rows)
    # Then one sheet per channel.
    for name, rows in channel_sheets:
        ws = wb.create_sheet(title=_safe_sheet_name(name, wb))
        _fill_sheet(ws, name, rows)
    return _save(wb)


def _safe_sheet_name(name, wb):
    """Excel sheet names: <=31 chars, no []:*?/\\, and must be unique."""
    bad = '[]:*?/\\'
    clean = "".join(("_" if ch in bad else ch) for ch in str(name)).strip()
    clean = (clean or "Sheet")[:31]
    base, n = clean, 2
    existing = set(wb.sheetnames)
    while clean in existing:
        suffix = f"_{n}"
        clean = base[:31 - len(suffix)] + suffix
        n += 1
    return clean


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    )


def _build_csv(show_rows):
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([display_label(c) for c in OUTPUT_COLS])
    prev_date = None
    for r in show_rows:
        d = r.get("Date", "")
        if prev_date is not None and d != prev_date:
            writer.writerow([])  # blank separator between dates
        writer.writerow([r.get(col, "") for col in OUTPUT_COLS])
        prev_date = d
    data = buf.getvalue().encode("utf-8-sig")
    return data, "text/csv", "csv"


def _fill_sheet(ws, title, show_rows):
    """Write the date-grouped Zee-format table into worksheet `ws`."""
    ws.title = _safe_sheet_name(title, ws.parent) \
        if ws.title != title else title

    amber = PatternFill("solid", start_color=HEADER_AMBER)
    seg_fill = PatternFill("solid", start_color=SEGMENT_BLUE)
    hdr_font = Font(name="Arial", bold=True, size=10, color="000000")
    cell_font = Font(name="Arial", size=10)
    wkcode_font = Font(name="Arial", bold=True, size=10, color="000000")
    sep_font = Font(name="Arial", bold=True, size=10, color="000000")
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    n_show = len(SHOW_INFO_COLS)        # 13
    n_break = len(BREAK_COLS)           # 36

    # --- Row 1: "Break N" group header over each Ad ST/ET pair --------------
    for i in range(18):
        col = n_show + 1 + i * 2        # first col of the pair (1-based)
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col + 1)
        c = ws.cell(row=1, column=col, value=f"Break {i + 1}")
        c.font = hdr_font
        c.fill = amber
        c.alignment = center
        c.border = border

    # --- Row 2: the 68 column headers --------------------------------------
    for ci, col in enumerate(OUTPUT_COLS, start=1):
        c = ws.cell(row=2, column=ci, value=display_label(col))
        c.font = hdr_font
        c.fill = amber
        c.alignment = center
        c.border = border

    ws.freeze_panes = "A3"

    wkcode_idx = 1                                  # WK CODE is column 1
    minsmiss_idx = OUTPUT_COLS.index("Mins Missing") + 1
    seg_start_idx = n_show + n_break + 1            # first segment column
    seg_indices = set(range(seg_start_idx, seg_start_idx + len(SEG_COLS)))

    # --- Data rows, grouped by date ----------------------------------------
    excel_row = 3
    prev_date = None
    for r in show_rows:
        d = r.get("Date", "")
        if prev_date is not None and d != prev_date:
            # blank separator row showing the WK CODE of the group below
            sep = ws.cell(row=excel_row, column=wkcode_idx,
                          value=r.get("WK CODE", ""))
            sep.font = sep_font
            # Make the segment cells on the separator row identical in
            # format to ordinary segment cells: same blue fill, same
            # border and font — only the value is blank.
            for ci in seg_indices:
                sc = ws.cell(row=excel_row, column=ci)
                sc.fill = seg_fill
                sc.border = border
                sc.font = cell_font
            excel_row += 1
        for ci, col in enumerate(OUTPUT_COLS, start=1):
            val = r.get(col, "")
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.font = cell_font
            c.border = border
            if ci == wkcode_idx:
                c.font = wkcode_font
            if ci in seg_indices:
                c.fill = seg_fill
            if ci == minsmiss_idx and val not in ("", None):
                fill = _mins_missing_fill(val)
                if fill:
                    c.fill = fill
        excel_row += 1
        prev_date = d

    # --- Column widths ------------------------------------------------------
    for ci, col in enumerate(OUTPUT_COLS, start=1):
        letter = ws.cell(row=2, column=ci).column_letter
        if col == "PROGNAME":
            ws.column_dimensions[letter].width = 28
        elif col in ("WK CODE", "Week", "Date", "Day", "Start Time"):
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 10

    return ws